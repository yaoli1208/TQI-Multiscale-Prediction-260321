#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Trident框架批量实验脚本
对iic_tqi_all.xlsx中的1203个样本进行批量实验
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
import warnings
import os
warnings.filterwarnings('ignore')

# 创建输出目录
os.makedirs('results/batch_experiments', exist_ok=True)

def load_data():
    """加载全量数据"""
    print("[1/3] 加载全量数据...")
    
    # 读取样本列表
    sample_df = pd.read_csv('data/suitable_samples.csv')
    sample_list = sample_df['tqi_mile'].tolist()  # 全部样本
    print(f"样本数量: {len(sample_list)}")
    
    # 读取原始数据
    print("[2/3] 读取原始检测数据...")
    wb = load_workbook('iic_tqi_all.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    all_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        all_data.append(row)
        if i % 50000 == 0 and i > 0:
            print(f"  已读取 {i:,} 行...")
    
    wb.close()
    
    # 转换为DataFrame
    df_all = pd.DataFrame(all_data, columns=['dete_dt', 'tqi_mile', 'tqi_val', 'tqi_lprf', 'tqi_rprf', 
                                              'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl'])
    df_all['tqi_mile'] = pd.to_numeric(df_all['tqi_mile'], errors='coerce')
    df_all['tqi_val'] = pd.to_numeric(df_all['tqi_val'], errors='coerce')
    df_all['dete_dt'] = pd.to_datetime(df_all['dete_dt'], errors='coerce')
    
    print(f"[3/3] 数据加载完成: {len(df_all):,}条记录")
    return df_all, sample_list

def clean_data(df):
    """数据清洗"""
    # 删除缺失值
    df = df.dropna(subset=['dete_dt', 'tqi_val'])
    
    if len(df) < 10:
        return df
    
    # IQR异常值检测
    Q1 = df['tqi_val'].quantile(0.25)
    Q3 = df['tqi_val'].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    df = df[(df['tqi_val'] >= lower_bound) & (df['tqi_val'] <= upper_bound)]
    
    # 按时间排序
    df = df.sort_values('dete_dt').reset_index(drop=True)
    
    return df

def calculate_seasonal(df):
    """计算季节性调整系数"""
    df['month'] = df['dete_dt'].dt.month
    seasonal_mean = df.groupby('month')['tqi_val'].mean()
    overall_mean = df['tqi_val'].mean()
    seasonal_adj = seasonal_mean - overall_mean
    return seasonal_adj, overall_mean

def run_experiments(df, tqi_mile):
    """对单个样本运行三种预测策略"""
    results = {
        'tqi_mile': tqi_mile,
        'original_count': len(df),
        'clean_count': 0,
        'time_span_days': 0,
        'tqi_mean': 0,
        'tqi_std': 0,
        'seasonal_amplitude': 0,
        'data_driven_mae': np.nan,
        'rolling_anchor_mae': np.nan,
        'post_maintenance_mae': np.nan,
        'best_strategy': '',
        'trident_improvement': np.nan,
        'is_trident_effective': False
    }
    
    # 数据清洗
    df_clean = clean_data(df)
    if len(df_clean) < 30:  # 数据量不足
        return results
    
    results['clean_count'] = len(df_clean)
    results['time_span_days'] = (df_clean['dete_dt'].max() - df_clean['dete_dt'].min()).days
    results['tqi_mean'] = df_clean['tqi_val'].mean()
    results['tqi_std'] = df_clean['tqi_val'].std()
    
    # 数据划分
    train_ratio = 0.7
    train_end = int(len(df_clean) * train_ratio)
    train_df = df_clean.iloc[:train_end].copy()
    test_df = df_clean.iloc[train_end:].copy()
    
    if len(test_df) < 10:
        return results
    
    # 季节性分析
    seasonal_adj, train_mean = calculate_seasonal(train_df)
    results['seasonal_amplitude'] = seasonal_adj.max() - seasonal_adj.min()
    
    y_true = test_df['tqi_val'].values
    
    # 策略1: 数据驱动（历史均值）
    pred_dd = np.full_like(y_true, train_mean)
    mae_dd = np.mean(np.abs(y_true - pred_dd))
    results['data_driven_mae'] = mae_dd
    
    # 策略2: 滚动锚定（最近年度均值 + 季节性）
    train_df['year'] = train_df['dete_dt'].dt.year
    yearly_mean = train_df.groupby('year')['tqi_val'].mean()
    recent_mean = yearly_mean.tail(3).mean() if len(yearly_mean) >= 3 else yearly_mean.mean()
    
    pred_ra = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        pred_ra.append(recent_mean + seasonal)
    mae_ra = np.mean(np.abs(y_true - np.array(pred_ra)))
    results['rolling_anchor_mae'] = mae_ra
    
    # 策略3: 修后预测（使用最近低点作为锚定）
    tqi_vals = train_df['tqi_val'].values
    local_mins = []
    for i in range(1, len(tqi_vals)-1):
        if tqi_vals[i] < tqi_vals[i-1] and tqi_vals[i] < tqi_vals[i+1]:
            if tqi_vals[i-1] - tqi_vals[i] > 0.3:
                local_mins.append(tqi_vals[i])
    
    if local_mins:
        anchor_val = local_mins[-1]
    else:
        anchor_val = train_mean
    
    pred_pm = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        pred_pm.append(anchor_val + seasonal)
    mae_pm = np.mean(np.abs(y_true - np.array(pred_pm)))
    results['post_maintenance_mae'] = mae_pm
    
    # 确定最佳策略
    maes = {
        'data_driven': mae_dd,
        'rolling_anchor': mae_ra,
        'post_maintenance': mae_pm
    }
    best_strategy = min(maes.items(), key=lambda x: x[1])
    results['best_strategy'] = best_strategy[0]
    
    # 判断Trident是否有效（滚动锚定或修后预测优于数据驱动）
    best_trident_mae = min(mae_ra, mae_pm)
    improvement = (mae_dd - best_trident_mae) / mae_dd * 100 if mae_dd > 0 else 0
    results['trident_improvement'] = improvement
    results['is_trident_effective'] = improvement > 10  # 提升>10%认为有效
    
    return results

def main():
    """主函数"""
    print("="*70)
    print("Trident框架批量实验")
    print("="*70)
    
    # 加载数据
    df_all, sample_list = load_data()
    
    # 批量实验
    print(f"\n[4/4] 开始批量实验 ({len(sample_list)}个样本)...")
    all_results = []
    
    for i, tqi_mile in enumerate(sample_list):
        # 提取单个样本数据
        df_sample = df_all[df_all['tqi_mile'] == tqi_mile].copy()
        
        if len(df_sample) < 50:
            continue
        
        # 运行实验
        result = run_experiments(df_sample, tqi_mile)
        all_results.append(result)
        
        if (i + 1) % 10 == 0 or i == len(sample_list) - 1:
            effective_count = sum([r['is_trident_effective'] for r in all_results])
            print(f"  进度: {i+1}/{len(sample_list)} | Trident有效: {effective_count}/{len(all_results)}")
    
    # 保存结果
    results_df = pd.DataFrame(all_results)
    results_df.to_csv('results/batch_experiments/all_samples_results.csv', index=False)
    print(f"\n实验结果已保存: results/batch_experiments/all_samples_results.csv")
    
    # 汇总统计
    print("\n" + "="*70)
    print("实验汇总")
    print("="*70)
    
    total = len(all_results)
    effective = sum([r['is_trident_effective'] for r in all_results])
    
    print(f"\n总样本数: {total}")
    print(f"Trident有效样本: {effective} ({effective/total*100:.1f}%)")
    print(f"Trident无效样本: {total-effective} ({(total-effective)/total*100:.1f}%)")
    
    # 有效样本特征
    if effective > 0:
        effective_samples = [r for r in all_results if r['is_trident_effective']]
        print(f"\n有效样本特征:")
        print(f"  平均记录数: {np.mean([r['clean_count'] for r in effective_samples]):.0f}")
        print(f"  平均时间跨度: {np.mean([r['time_span_days'] for r in effective_samples]):.0f}天")
        print(f"  平均TQI均值: {np.mean([r['tqi_mean'] for r in effective_samples]):.3f}")
        print(f"  平均季节性振幅: {np.mean([r['seasonal_amplitude'] for r in effective_samples]):.3f}")
        print(f"  平均提升幅度: {np.mean([r['trident_improvement'] for r in effective_samples]):.1f}%")
    
    # 保存有效样本列表
    effective_df = results_df[results_df['is_trident_effective'] == True]
    effective_df.to_csv('results/batch_experiments/effective_samples.csv', index=False)
    print(f"\n有效样本列表已保存: results/batch_experiments/effective_samples.csv")
    
    print("\n实验完成!")

if __name__ == '__main__':
    main()

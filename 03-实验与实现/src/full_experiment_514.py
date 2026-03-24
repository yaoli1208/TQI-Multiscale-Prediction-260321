#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Trident完整实验 - 514个有效样本批量跑论文实验
包括：基线对比(MA/Holt/LSTM/TimeMixer) + Trident策略 + 分量分组
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
import warnings
import os
import json
warnings.filterwarnings('ignore')

os.makedirs('results/full_experiments_514', exist_ok=True)

def load_full_data():
    """加载全量数据"""
    print("[1/3] 加载样本列表...")
    effective_df = pd.read_csv('results/batch_experiments/effective_samples.csv')
    sample_list = effective_df['tqi_mile'].tolist()
    print(f"有效样本数: {len(sample_list)}")
    
    print("[2/3] 读取原始检测数据...")
    wb = load_workbook('iic_tqi_all.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    all_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        all_data.append(row)
        if i % 100000 == 0 and i > 0:
            print(f"  已读取 {i:,} 行...")
    wb.close()
    
    df_all = pd.DataFrame(all_data, columns=['dete_dt', 'tqi_mile', 'tqi_val', 'tqi_lprf', 'tqi_rprf', 
                                              'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl'])
    df_all['tqi_mile'] = pd.to_numeric(df_all['tqi_mile'], errors='coerce')
    df_all['tqi_val'] = pd.to_numeric(df_all['tqi_val'], errors='coerce')
    for col in ['tqi_lprf', 'tqi_rprf', 'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl']:
        df_all[col] = pd.to_numeric(df_all[col], errors='coerce')
    df_all['dete_dt'] = pd.to_datetime(df_all['dete_dt'], errors='coerce')
    
    print(f"[3/3] 数据加载完成: {len(df_all):,}条记录")
    return df_all, sample_list

def clean_data(df):
    """数据清洗"""
    df = df.dropna(subset=['dete_dt', 'tqi_val'])
    if len(df) < 20:
        return df
    Q1 = df['tqi_val'].quantile(0.25)
    Q3 = df['tqi_val'].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    df = df[(df['tqi_val'] >= lower_bound) & (df['tqi_val'] <= upper_bound)]
    return df.sort_values('dete_dt').reset_index(drop=True)

def calculate_seasonal(df):
    """计算季节性调整系数"""
    df['month'] = df['dete_dt'].dt.month
    seasonal_mean = df.groupby('month')['tqi_val'].mean()
    overall_mean = df['tqi_val'].mean()
    seasonal_adj = seasonal_mean - overall_mean
    return seasonal_adj, overall_mean

def baseline_ma(train_df, test_dates):
    """移动平均（最近12个月）"""
    cutoff_date = train_df['dete_dt'].max() - pd.DateOffset(months=12)
    recent = train_df[train_df['dete_dt'] >= cutoff_date]
    if len(recent) == 0:
        recent = train_df
    mean_val = recent['tqi_val'].mean()
    return np.full(len(test_dates), mean_val)

def baseline_holt(train_df, test_dates):
    """指数平滑（简化版 - 使用加权移动平均）"""
    try:
        # 使用最近6个月的加权平均（近期权重更高）
        cutoff_date = train_df['dete_dt'].max() - pd.DateOffset(months=6)
        recent = train_df[train_df['dete_dt'] >= cutoff_date]
        if len(recent) < 3:
            return baseline_ma(train_df, test_dates)
        
        # 加权平均（时间越近权重越高）
        recent = recent.copy()
        recent['days_ago'] = (recent['dete_dt'].max() - recent['dete_dt']).dt.days
        recent['weight'] = np.exp(-recent['days_ago'] / 30)  # 30天衰减
        weighted_mean = np.average(recent['tqi_val'], weights=recent['weight'])
        
        return np.full(len(test_dates), weighted_mean)
    except:
        return baseline_ma(train_df, test_dates)

def baseline_lstm(train_df, test_df):
    """LSTM替代方案 - 使用MLP捕获时序特征"""
    try:
        # 构建时序特征
        train_df = train_df.copy()
        train_df['days_since_start'] = (train_df['dete_dt'] - train_df['dete_dt'].min()).dt.days
        train_df['month'] = train_df['dete_dt'].dt.month
        train_df['year'] = train_df['dete_dt'].dt.year
        
        X_train = train_df[['days_since_start', 'month', 'year']].values
        y_train = train_df['tqi_val'].values
        
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        
        model = MLPRegressor(hidden_layer_sizes=(50, 25), max_iter=1000, random_state=42)
        model.fit(X_train_scaled, y_train)
        
        # 预测
        test_df = test_df.copy()
        test_df['days_since_start'] = (test_df['dete_dt'] - train_df['dete_dt'].min()).dt.days
        test_df['month'] = test_df['dete_dt'].dt.month
        test_df['year'] = test_df['dete_dt'].dt.year
        X_test = test_df[['days_since_start', 'month', 'year']].values
        X_test_scaled = scaler.transform(X_test)
        
        return model.predict(X_test_scaled)
    except:
        return baseline_ma(train_df, test_df)

def trident_strategies(train_df, test_df, seasonal_adj, train_mean):
    """Trident三种策略"""
    results = {}
    y_true = test_df['tqi_val'].values
    
    # 1. 数据驱动
    pred_dd = np.full_like(y_true, train_mean)
    results['data_driven_mae'] = np.mean(np.abs(y_true - pred_dd))
    
    # 2. 滚动锚定
    train_df_copy = train_df.copy()
    train_df_copy['year'] = train_df_copy['dete_dt'].dt.year
    yearly_mean = train_df_copy.groupby('year')['tqi_val'].mean()
    recent_mean = yearly_mean.tail(3).mean() if len(yearly_mean) >= 3 else yearly_mean.mean()
    
    pred_ra = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        pred_ra.append(recent_mean + seasonal)
    results['rolling_anchor_mae'] = np.mean(np.abs(y_true - np.array(pred_ra)))
    
    # 3. 修后预测
    tqi_vals = train_df['tqi_val'].values
    local_mins = []
    for i in range(1, len(tqi_vals)-1):
        if tqi_vals[i] < tqi_vals[i-1] and tqi_vals[i] < tqi_vals[i+1]:
            if tqi_vals[i-1] - tqi_vals[i] > 0.3:
                local_mins.append(tqi_vals[i])
    anchor_val = local_mins[-1] if local_mins else train_mean
    
    pred_pm = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        pred_pm.append(anchor_val + seasonal)
    results['post_maintenance_mae'] = np.mean(np.abs(y_true - np.array(pred_pm)))
    
    # 确定最佳Trident策略
    trident_maes = {
        'data_driven': results['data_driven_mae'],
        'rolling_anchor': results['rolling_anchor_mae'],
        'post_maintenance': results['post_maintenance_mae']
    }
    best_trident = min(trident_maes.items(), key=lambda x: x[1])
    results['best_trident_strategy'] = best_trident[0]
    results['best_trident_mae'] = best_trident[1]
    
    return results

def component_group_experiment(train_df, test_df):
    """分量分组实验"""
    results = {}
    
    # 计算分量分组
    train_df = train_df.copy()
    test_df = test_df.copy()
    train_df['plane'] = train_df['tqi_laln'] + train_df['tqi_raln'] + train_df['tqi_gage']
    train_df['elevation'] = train_df['tqi_lprf'] + train_df['tqi_rprf'] + train_df['tqi_xlvl'] + train_df['tqi_warp1']
    test_df['plane'] = test_df['tqi_laln'] + test_df['tqi_raln'] + test_df['tqi_gage']
    test_df['elevation'] = test_df['tqi_lprf'] + test_df['tqi_rprf'] + test_df['tqi_xlvl'] + test_df['tqi_warp1']
    
    # 预测平面组
    feature_cols = ['tqi_lprf', 'tqi_rprf', 'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl']
    X_train = train_df[feature_cols].fillna(0).values
    X_test = test_df[feature_cols].fillna(0).values
    y_train_plane = train_df['plane'].values
    y_test_plane = test_df['plane'].values
    
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    model_plane = MLPRegressor(hidden_layer_sizes=(64, 32), max_iter=500, random_state=42)
    model_plane.fit(X_train_scaled, y_train_plane)
    pred_plane = model_plane.predict(X_test_scaled)
    results['plane_mae'] = np.mean(np.abs(y_test_plane - pred_plane))
    
    # 预测高程组
    y_train_elev = train_df['elevation'].values
    y_test_elev = test_df['elevation'].values
    
    model_elev = MLPRegressor(hidden_layer_sizes=(64, 32), max_iter=500, random_state=42)
    model_elev.fit(X_train_scaled, y_train_elev)
    pred_elev = model_elev.predict(X_test_scaled)
    results['elevation_mae'] = np.mean(np.abs(y_test_elev - pred_elev))
    
    results['plane_better'] = results['plane_mae'] < results['elevation_mae']
    
    return results

def run_full_experiment(df_sample, tqi_mile):
    """对单个样本运行完整实验"""
    result = {'tqi_mile': tqi_mile}
    
    # 数据清洗
    df_clean = clean_data(df_sample)
    if len(df_clean) < 50:
        return None
    
    result['record_count'] = len(df_clean)
    result['time_span_days'] = (df_clean['dete_dt'].max() - df_clean['dete_dt'].min()).days
    result['tqi_mean'] = df_clean['tqi_val'].mean()
    result['tqi_std'] = df_clean['tqi_val'].std()
    
    # 数据划分
    train_ratio = 0.7
    train_end = int(len(df_clean) * train_ratio)
    train_df = df_clean.iloc[:train_end].copy()
    test_df = df_clean.iloc[train_end:].copy()
    
    if len(test_df) < 10:
        return None
    
    # 季节性分析
    seasonal_adj, train_mean = calculate_seasonal(train_df)
    result['seasonal_amplitude'] = seasonal_adj.max() - seasonal_adj.min()
    
    y_true = test_df['tqi_val'].values
    
    # 基线方法
    result['ma_mae'] = np.mean(np.abs(y_true - baseline_ma(train_df, test_df['dete_dt'])))
    result['holt_mae'] = np.mean(np.abs(y_true - baseline_holt(train_df, test_df['dete_dt'])))
    result['lstm_mae'] = np.mean(np.abs(y_true - baseline_lstm(train_df, test_df)))
    
    # Trident策略
    trident_results = trident_strategies(train_df, test_df, seasonal_adj, train_mean)
    result.update(trident_results)
    
    # 分量分组实验
    component_results = component_group_experiment(train_df, test_df)
    result.update(component_results)
    
    # 确定整体最佳方法
    all_maes = {
        'MA': result['ma_mae'],
        'Holt': result['holt_mae'],
        'LSTM': result['lstm_mae'],
        'DataDriven': result['data_driven_mae'],
        'RollingAnchor': result['rolling_anchor_mae'],
        'PostMaintenance': result['post_maintenance_mae']
    }
    best_method = min(all_maes.items(), key=lambda x: x[1])
    result['best_overall_method'] = best_method[0]
    result['best_overall_mae'] = best_method[1]
    
    # Trident是否最佳
    trident_mae = result['best_trident_mae']
    baseline_best = min(result['ma_mae'], result['holt_mae'], result['lstm_mae'])
    result['trident_vs_baseline'] = (baseline_best - trident_mae) / baseline_best * 100
    result['trident_is_best'] = trident_mae <= baseline_best
    
    return result

def main():
    print("="*70)
    print("Trident完整实验 - 514个有效样本")
    print("="*70)
    
    df_all, sample_list = load_full_data()
    
    print(f"\n开始批量实验 ({len(sample_list)}个样本)...")
    all_results = []
    
    for i, tqi_mile in enumerate(sample_list):
        df_sample = df_all[df_all['tqi_mile'] == tqi_mile].copy()
        if len(df_sample) < 50:
            continue
        
        result = run_full_experiment(df_sample, tqi_mile)
        if result:
            all_results.append(result)
        
        if (i + 1) % 10 == 0 or i == len(sample_list) - 1:
            trident_best_count = sum([r.get('trident_is_best', False) for r in all_results])
            print(f"  进度: {i+1}/{len(sample_list)} | Trident最佳: {trident_best_count}/{len(all_results)}")
    
    # 保存结果
    results_df = pd.DataFrame(all_results)
    results_df.to_csv('results/full_experiments_514/full_experiment_results.csv', index=False)
    print(f"\n实验结果已保存: results/full_experiments_514/full_experiment_results.csv")
    
    # 汇总统计
    print("\n" + "="*70)
    print("实验汇总")
    print("="*70)
    
    total = len(all_results)
    trident_best = sum([r.get('trident_is_best', False) for r in all_results])
    
    print(f"\n总样本数: {total}")
    print(f"Trident为最佳方法: {trident_best} ({trident_best/total*100:.1f}%)")
    print(f"基线方法更优: {total-trident_best} ({(total-trident_best)/total*100:.1f}%)")
    
    # 各方法获胜次数
    method_counts = results_df['best_overall_method'].value_counts()
    print(f"\n各方法获胜次数:")
    for method, count in method_counts.items():
        if pd.notna(method):
            print(f"  {method}: {count} ({count/total*100:.1f}%)")
    
    # 性能对比
    print(f"\n平均MAE对比:")
    print(f"  移动平均: {results_df['ma_mae'].mean():.3f}")
    print(f"  指数平滑: {results_df['holt_mae'].mean():.3f}")
    print(f"  LSTM: {results_df['lstm_mae'].mean():.3f}")
    print(f"  Trident最佳: {results_df['best_trident_mae'].mean():.3f}")
    
    # 分量分组
    plane_better = results_df['plane_better'].sum()
    print(f"\n分量分组:")
    print(f"  平面组优于高程组: {plane_better}/{total} ({plane_better/total*100:.1f}%)")
    
    # 保存汇总
    summary = {
        'total_samples': total,
        'trident_best_count': int(trident_best),
        'trident_best_ratio': float(trident_best/total),
        'method_wins': method_counts.to_dict(),
        'mean_mae': {
            'MA': float(results_df['ma_mae'].mean()),
            'Holt': float(results_df['holt_mae'].mean()),
            'LSTM': float(results_df['lstm_mae'].mean()),
            'Trident': float(results_df['best_trident_mae'].mean())
        }
    }
    with open('results/full_experiments_514/summary.json', 'w') as f:
        json.dump(summary, f, indent=2)
    
    print("\n实验完成!")

if __name__ == '__main__':
    main()

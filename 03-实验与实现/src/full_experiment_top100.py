#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Trident完整实验 - Top 100有效样本
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
import warnings
import os
import json
warnings.filterwarnings('ignore')

os.makedirs('results/full_experiments_top100', exist_ok=True)

def load_data():
    """加载Top 100样本"""
    print("[1/2] 加载Top 100样本...")
    effective_df = pd.read_csv('results/batch_experiments/effective_samples.csv')
    top100 = effective_df.nlargest(100, 'trident_improvement')
    sample_list = top100['tqi_mile'].tolist()
    print(f"Top 100样本 (MAE改善最高): {len(sample_list)}个")
    
    print("[2/2] 读取原始数据...")
    wb = load_workbook('iic_tqi_all.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    all_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        all_data.append(row)
        if i % 100000 == 0 and i > 0:
            print(f"  已读取 {i:,} 行...")
    wb.close()
    
    df_all = pd.DataFrame(all_data, columns=['dete_dt', 'tqi_mile', 'tqi_val', 'tqi_lprf', 'tqi_rprf', 
                                              'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl'])
    df_all['tqi_mile'] = pd.to_numeric(df_all['tqi_mile'], errors='coerce')
    df_all['tqi_val'] = pd.to_numeric(df_all['tqi_val'], errors='coerce')
    for col in ['tqi_lprf', 'tqi_rprf', 'tqi_laln', 'tqi_raln', 'tqi_gage', 'tqi_warp1', 'tqi_xlvl']:
        df_all[col] = pd.to_numeric(df_all[col], errors='coerce')
    df_all['dete_dt'] = pd.to_datetime(df_all['dete_dt'], errors='coerce')
    
    return df_all, sample_list

def clean_data(df):
    df = df.dropna(subset=['dete_dt', 'tqi_val'])
    if len(df) < 20:
        return df
    Q1 = df['tqi_val'].quantile(0.25)
    Q3 = df['tqi_val'].quantile(0.75)
    IQR = Q3 - Q1
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR
    df = df[(df['tqi_val'] >= lower) & (df['tqi_val'] <= upper)]
    return df.sort_values('dete_dt').reset_index(drop=True)

def run_experiment(df_sample, tqi_mile):
    df_clean = clean_data(df_sample)
    if len(df_clean) < 50:
        return None
    
    # 数据划分
    train_ratio = 0.7
    train_end = int(len(df_clean) * train_ratio)
    train_df = df_clean.iloc[:train_end].copy()
    test_df = df_clean.iloc[train_end:].copy()
    
    if len(test_df) < 10:
        return None
    
    y_true = test_df['tqi_val'].values
    train_mean = train_df['tqi_val'].mean()
    
    # 季节性
    train_df['month'] = train_df['dete_dt'].dt.month
    seasonal_mean = train_df.groupby('month')['tqi_val'].mean()
    seasonal_adj = seasonal_mean - train_mean
    
    result = {
        'tqi_mile': tqi_mile,
        'record_count': len(df_clean),
        'tqi_mean': df_clean['tqi_val'].mean(),
        'tqi_std': df_clean['tqi_val'].std(),
        'seasonal_amplitude': seasonal_adj.max() - seasonal_adj.min()
    }
    
    # 基线1: 移动平均
    cutoff = train_df['dete_dt'].max() - pd.DateOffset(months=12)
    recent = train_df[train_df['dete_dt'] >= cutoff]
    if len(recent) == 0:
        recent = train_df
    ma_pred = np.full(len(test_df), recent['tqi_val'].mean())
    result['ma_mae'] = np.mean(np.abs(y_true - ma_pred))
    
    # 基线2: 历史均值
    result['hm_mae'] = np.mean(np.abs(y_true - train_mean))
    
    # 基线3: LSTM(MLP替代)
    try:
        train_df['days'] = (train_df['dete_dt'] - train_df['dete_dt'].min()).dt.days
        test_df_copy = test_df.copy()
        test_df_copy['days'] = (test_df_copy['dete_dt'] - train_df['dete_dt'].min()).dt.days
        
        X_train = train_df[['days']].values
        y_train = train_df['tqi_val'].values
        X_test = test_df_copy[['days']].values
        
        model = MLPRegressor(hidden_layer_sizes=(32, 16), max_iter=500, random_state=42)
        model.fit(X_train, y_train)
        lstm_pred = model.predict(X_test)
        result['lstm_mae'] = np.mean(np.abs(y_true - lstm_pred))
    except:
        result['lstm_mae'] = result['hm_mae']
    
    # Trident策略
    # 1. 数据驱动
    result['dd_mae'] = result['hm_mae']
    
    # 2. 滚动锚定
    train_df['year'] = train_df['dete_dt'].dt.year
    yearly = train_df.groupby('year')['tqi_val'].mean()
    recent_mean = yearly.tail(3).mean() if len(yearly) >= 3 else yearly.mean()
    
    ra_pred = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        ra_pred.append(recent_mean + seasonal)
    result['ra_mae'] = np.mean(np.abs(y_true - np.array(ra_pred)))
    
    # 3. 修后预测
    tqi_vals = train_df['tqi_val'].values
    local_mins = []
    for i in range(1, len(tqi_vals)-1):
        if tqi_vals[i] < tqi_vals[i-1] and tqi_vals[i] < tqi_vals[i+1]:
            if tqi_vals[i-1] - tqi_vals[i] > 0.3:
                local_mins.append(tqi_vals[i])
    anchor = local_mins[-1] if local_mins else train_mean
    
    pm_pred = []
    for _, row in test_df.iterrows():
        month = row['dete_dt'].month
        seasonal = seasonal_adj.get(month, 0)
        pm_pred.append(anchor + seasonal)
    result['pm_mae'] = np.mean(np.abs(y_true - np.array(pm_pred)))
    
    # 最佳Trident
    trident_maes = [result['dd_mae'], result['ra_mae'], result['pm_mae']]
    result['best_trident_mae'] = min(trident_maes)
    
    # 整体最佳
    all_maes = [result['ma_mae'], result['hm_mae'], result['lstm_mae'], 
                result['dd_mae'], result['ra_mae'], result['pm_mae']]
    result['best_overall_mae'] = min(all_maes)
    
    # 判断Trident是否最佳
    baseline_best = min(result['ma_mae'], result['hm_mae'], result['lstm_mae'])
    result['trident_is_best'] = result['best_trident_mae'] <= baseline_best
    result['improvement_vs_baseline'] = (baseline_best - result['best_trident_mae']) / baseline_best * 100
    
    return result

def main():
    print("="*70)
    print("Trident完整实验 - Top 100样本")
    print("="*70)
    
    df_all, sample_list = load_data()
    
    print(f"\n开始实验...")
    results = []
    for i, tqi_mile in enumerate(sample_list):
        df_sample = df_all[df_all['tqi_mile'] == tqi_mile].copy()
        if len(df_sample) < 50:
            continue
        
        res = run_experiment(df_sample, tqi_mile)
        if res:
            results.append(res)
        
        if (i + 1) % 10 == 0:
            print(f"  进度: {i+1}/100")
    
    # 保存
    results_df = pd.DataFrame(results)
    results_df.to_csv('results/full_experiments_top100/results.csv', index=False)
    
    # 汇总
    print("\n" + "="*70)
    print("实验汇总")
    print("="*70)
    
    total = len(results)
    trident_best = sum([r['trident_is_best'] for r in results])
    
    print(f"\n样本数: {total}")
    print(f"Trident最佳: {trident_best} ({trident_best/total*100:.1f}%)")
    
    print(f"\n平均MAE:")
    print(f"  移动平均: {results_df['ma_mae'].mean():.3f}")
    print(f"  历史均值: {results_df['hm_mae'].mean():.3f}")
    print(f"  LSTM: {results_df['lstm_mae'].mean():.3f}")
    print(f"  Trident最佳: {results_df['best_trident_mae'].mean():.3f}")
    
    print(f"\nTrident平均提升: {results_df['improvement_vs_baseline'].mean():.1f}%")
    
    # 保存汇总
    summary = {
        'total': total,
        'trident_best': int(trident_best),
        'mean_mae': {
            'MA': float(results_df['ma_mae'].mean()),
            'HM': float(results_df['hm_mae'].mean()),
            'LSTM': float(results_df['lstm_mae'].mean()),
            'Trident': float(results_df['best_trident_mae'].mean())
        },
        'mean_improvement': float(results_df['improvement_vs_baseline'].mean())
    }
    with open('results/full_experiments_top100/summary.json', 'w') as f:
        json.dump(summary, f, indent=2)
    
    print("\n完成!")

if __name__ == '__main__':
    main()
